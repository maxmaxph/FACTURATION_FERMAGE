import os
import sys
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk  # Pour les listes déroulantes
from openpyxl import load_workbook  # type: ignore
from fpdf import FPDF  # type: ignore
from datetime import date
from PIL import Image, ImageTk  # Pour afficher le logo
import webbrowser
import subprocess  # Ajouter cette importation en haut de votre script

# Obtenir le chemin du répertoire du script pour les fichiers relatifs
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE_PATH = os.path.join(BASE_DIR, 'tableau/tableau_fermage.xlsx')
FONTS_DIR = os.path.join(BASE_DIR, 'assets/fonts')
LOGO_PATH = os.path.join(BASE_DIR, 'assets/img/Fermage.png')
ICON_PATH = os.path.join(BASE_DIR, 'assets/img/facturation_fermage.ico')
# Définir le chemin de sortie pour les factures dans le même répertoire que l'exécutable
PDF_OUTPUT_DIR = os.path.join(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else BASE_DIR), 'factures')
# Vérification et création du dossier de factures si non existant
if not os.path.exists(PDF_OUTPUT_DIR):
    os.makedirs(PDF_OUTPUT_DIR)

# Définir les constantes d’interface
APP_TITLE = "Générateur de Facture de Fermage"
APP_GEOMETRY = "500x650"
BACKGROUND_COLOR = "#f0f0f0"
BUTTON_COLOR = "#4CAF50"
BUTTON_TEXT_COLOR = "white"
TITLE_FONT = ("Arial", 11, "bold")
LABEL_FONT = ("Arial", 8)
BUTTON_FONT = ("Arial", 10, "bold")
DEFAULT_YEAR_OPTIONS = ["2023", "2024", "2025", "2026", "2027", "2028"]
DEFAULT_OPTION = "Sélectionner"

# Constantes pour le PDF
PDF_HEADER_FONT = "DejaVuSans"
PDF_HEADER_FONT_SIZE = 14
PDF_TEXT_FONT_SIZE = 10
PDF_FOOTER_FONT_SIZE = 8

# Auteur et informations de copyright
AUTHOR_NAME = "Maxime LENFANT"
AUTHOR_WEBSITE = "https://maxime-lenfant.fr"
COPYRIGHT_YEAR = "2024"
APP_VERSION = "1.1"


class PDF(FPDF):
    def __init__(self):
        super().__init__()
        self.add_font(PDF_HEADER_FONT, '', os.path.join(FONTS_DIR, 'DejaVuSans.ttf'), uni=True)
        self.add_font(PDF_HEADER_FONT, 'B', os.path.join(FONTS_DIR, 'DejaVuSans-Bold.ttf'), uni=True)
        self.add_font(PDF_HEADER_FONT, 'I', os.path.join(FONTS_DIR, 'DejaVuSans-Oblique.ttf'), uni=True)
        self.add_font(PDF_HEADER_FONT, 'BI', os.path.join(FONTS_DIR, 'DejaVuSans-BoldOblique.ttf'), uni=True)

    def header(self):
        self.set_font(PDF_HEADER_FONT, "B", PDF_HEADER_FONT_SIZE)
        self.cell(0, 10, "Facture de Fermage", ln=True, align="C")
        self.set_font(PDF_HEADER_FONT, "", PDF_TEXT_FONT_SIZE)
        self.cell(0, 10, f"Date : {date.today().strftime('%d/%m/%Y')}", ln=True, align="C")
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font(PDF_HEADER_FONT, "I", PDF_FOOTER_FONT_SIZE)
        self.cell(0, 10, "Page %s" % self.page_no(), 0, 0, 'C')

# Fonction pour ouvrir un dossier
def open_folder(folder_path):
    if sys.platform == "win32":  # Pour Windows
        os.startfile(folder_path)
    elif sys.platform == "darwin":  # Pour macOS
        subprocess.Popen(["open", folder_path])
    elif sys.platform == "linux":  # Pour Linux
        subprocess.Popen(["xdg-open", folder_path])

# Fonction pour charger les données pour une année spécifique
def charger_donnees(annee_selectionnee):
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        proprietaires = set()
        fermiers = set()
        parcelles_proprietaires = {}

        if annee_selectionnee not in wb.sheetnames:
            messagebox.showerror("Erreur", f"La feuille pour l'année {annee_selectionnee} n'existe pas.")
            return [], [], {}, {}

        ws = wb[annee_selectionnee]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # Validation et formatage de chaque champ
                proprietaire = str(row[1]) if row[1] is not None else ""
                fermier = str(row[5]) if row[5] is not None else ""
                parcelle = str(row[4]) if row[4] is not None else ""
                
                # Conversion de surface en float, avec gestion des erreurs
                try:
                    surface = float(row[8]) if row[8] is not None else 0.0
                except ValueError:
                    messagebox.showwarning("Avertissement", f"Valeur incorrecte pour la surface : {row[8]}. La surface sera définie à 0.")
                    surface = 0.0
                
                # Ajout des données seulement si elles sont complètes et correctes
                if proprietaire and fermier and parcelle and surface:
                    proprietaires.add(proprietaire)
                    fermiers.add(fermier)
                    parcelles_proprietaires.setdefault(proprietaire, []).append((parcelle, surface))
            
            except (IndexError, TypeError) as e:
                # Gestion des erreurs spécifiques
                print(f"Erreur lors de la lecture d'une ligne : {e}")
                messagebox.showerror("Erreur", "Certaines colonnes du fichier Excel contiennent des données incorrectes.")
                return [], [], {}, {}

        return list(proprietaires), list(fermiers), parcelles_proprietaires

    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible de charger les données : {e}")
        return [], [], {}, {}





# Fonction pour mettre à jour les parcelles
def update_parcelles():
    annee = combo_annee.get()
    proprietaire = combo_proprietaire.get()

    if not annee:
        messagebox.showwarning("Attention", "Veuillez sélectionner une année.")
        return
    if not proprietaire:
        messagebox.showwarning("Attention", "Veuillez sélectionner un propriétaire.")
        return

    _, _, parcelles_proprietaires = charger_donnees(annee)
    parcelles = parcelles_proprietaires.get(proprietaire, [])

    for widget in parcelles_frame.winfo_children():
        widget.destroy()

    global parcelle_vars
    parcelle_vars = []
    for parcelle, surface in parcelles:
        var = tk.BooleanVar()
        cb = tk.Checkbutton(parcelles_frame, text=f"{parcelle} - {surface} ha", variable=var)
        cb.var = var
        cb.grid(sticky="w")
        parcelle_vars.append((parcelle, surface, var))


# Fonction pour mettre à jour les propriétaires et fermiers selon l'année
def on_annee_change(event):
    annee = combo_annee.get()
    proprietaires, fermiers, _ = charger_donnees(annee)
    combo_proprietaire['values'] = [DEFAULT_OPTION] + proprietaires  # Ajouter l'option par défaut
    combo_proprietaire.set(DEFAULT_OPTION)  # Définir comme valeur initiale
    
    combo_fermier['values'] = [DEFAULT_OPTION] + fermiers  # Ajouter l'option par défaut
    combo_fermier.set(DEFAULT_OPTION)  # Définir comme valeur initiale
    
    for widget in parcelles_frame.winfo_children():
        widget.destroy()


def on_proprietaire_change(event):
    update_parcelles()


def generer_facture_pdf(annee, nom_proprietaire, nom_fermier, parcelles_selectionnees):
    # Définir le dossier de sauvegarde pour les fichiers PDF
    dossier_factures = PDF_OUTPUT_DIR  # Utilisation de la constante définie pour le dossier des factures

    # Créer le dossier `factures` s'il n'existe pas
    if not os.path.exists(dossier_factures):
        os.makedirs(dossier_factures)

    # Définir le chemin complet du fichier PDF à sauvegarder dans le dossier `factures`
    fichier_pdf = os.path.join(dossier_factures, f'facture_fermage_{annee}_{nom_proprietaire}_{nom_fermier}.pdf')

    try:
        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
        if annee not in wb.sheetnames:
            messagebox.showerror("Erreur", f"La feuille pour l'année {annee} n'existe pas.")
            return

        ws = wb[annee]

        adresse_proprietaire, cp_ville_proprietaire, adresse_fermier, cp_ville_fermier = "", "", "", ""
        indice_actuel = 1
        prix_quintal = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == nom_proprietaire and row[5] == nom_fermier:
                adresse_proprietaire = row[2] or "Adresse non spécifiée"
                cp_ville_proprietaire = row[3] or "CP/Ville non spécifiés"
                adresse_fermier = row[6] or "Adresse non spécifiée"
                cp_ville_fermier = row[7] or "CP/Ville non spécifiés"
                prix_quintal = float(row[10]) if row[10] else 0
                indice_actuel = float(row[11]) if row[11] else 1
                break
        else:
            messagebox.showerror("Erreur", "Informations de propriétaire/fermier introuvables.")
            return

        # Initialiser le PDF
        pdf = PDF()
        pdf.add_page()

        # En-tête et infos propriétaire/fermier
        pdf.set_font("DejaVuSans", "B", 10)
        pdf.cell(100, 8, "PROPRIÉTAIRE :", ln=False)
        pdf.cell(0, 8, "FERMIER :", ln=True)
        pdf.set_font("DejaVuSans", "", 10)
        pdf.cell(100, 8, f"{nom_proprietaire}", ln=False)
        pdf.cell(0, 8, f"{nom_fermier}", ln=True)
        pdf.cell(100, 8, f"{adresse_proprietaire}", ln=False)
        pdf.cell(0, 8, f"{adresse_fermier}", ln=True)
        pdf.cell(100, 8, f"{cp_ville_proprietaire}", ln=False)
        pdf.cell(0, 8, f"{cp_ville_fermier}", ln=True)
        pdf.ln(20)
        
        # Ville et date
        ville_proprietaire = cp_ville_proprietaire.split()[-1]
        pdf.cell(0, 8, f"{ville_proprietaire}, le {date.today().strftime('%d/%m/%Y')}", ln=True)
        pdf.ln(10)

        # Explication du tarif
        pdf.set_font("DejaVuSans", "", 10)
        parcelle_text = ", ".join([str(p[0]) for p in parcelles_selectionnees])
        surface_text = ", ".join([str(p[1]) for p in parcelles_selectionnees])
        pdf.write(10, f"{nom_fermier} doit le fermage de l'année {annee} pour les parcelles {parcelle_text} de surface respective {surface_text} ha.\n")
        pdf.write(10, "Le tarif est calculé par rapport au prix du quintal de référence : ")

        # Prix du quintal et indice d'ajustement en gras
        pdf.set_font("DejaVuSans", "B", 10)
        pdf.write(10, f"{prix_quintal:.2f} € ")
        pdf.set_font("DejaVuSans", "", 10)
        pdf.write(10, "ajusté par l'indice actuel d'ajustement de ")
        pdf.set_font("DejaVuSans", "B", 10)
        pdf.write(10, f"{indice_actuel}%\n")

        # Table de détails pour chaque parcelle
        pdf.set_font("DejaVuSans", "B", 10)
        headers = ["N° parcelle", "Surface (ha)", "Quantité (qx)", "Prix ajusté (€)", "Impôts (%)","Total HT (€)"]
        column_widths = [35, 30, 30, 30, 30, 30]

        for i, header in enumerate(headers):
            pdf.cell(column_widths[i], 10, header, border=1, align="C")
        pdf.ln(10)

        # Remplir les données du tableau
        pdf.set_font("DejaVuSans", "", 10)
        total_ht = 0
        total_ttc = 0
        for parcelle, surface in parcelles_selectionnees:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == nom_proprietaire and row[5] == nom_fermier and row[4] == parcelle:
                    quantite = row[9]  # Colonne J : Quantité en quintaux (qx)
                    prix_quintal_adjusted = float(row[12]) if row[12] else 0  # Colonne M : Prix ajusté du quintal (€)
                    taxe = row[13]  # Colonne N : Taxe et impôts (%)
                    
                    montant_ht = surface * quantite * prix_quintal_adjusted
                    montant_ttc = montant_ht * (1 + taxe / 100)
                    
                    pdf.cell(column_widths[0], 10, parcelle, border=1, align="C")
                    pdf.cell(column_widths[1], 10, f"{surface:.2f}", border=1, align="C")
                    pdf.cell(column_widths[2], 10, f"{quantite:.2f}", border=1, align="C")
                    pdf.cell(column_widths[3], 10, f"{prix_quintal_adjusted:.2f}", border=1, align="C")
                    pdf.cell(column_widths[4], 10, f"{taxe:.2f}", border=1, align="C")
                    pdf.cell(column_widths[5], 10, f"{montant_ht:.2f}", border=1, align="C")
                    
                    pdf.ln(10)
                    
                    total_ht += montant_ht
                    total_ttc += montant_ttc
                    break

        # Après la boucle de remplissage des données du tableau
        pdf.set_font("DejaVuSans", "B", 10)

       # Ligne pour Total HT sans bordure sur les cellules vides
        pdf.cell(sum(column_widths[:-2]), 10, "", border=0, align="C")  # Cellule vide sans bordure pour l'alignement
        pdf.cell(column_widths[-2], 10, "Total HT (€)", border=1, align="R")
        pdf.cell(column_widths[-1], 10, f"{total_ht:.2f}", border=1, align="C")
        pdf.ln(10)

        # Ligne pour Total TTC sans bordure sur les cellules vides
        pdf.cell(sum(column_widths[:-2]), 10, "", border=0, align="C")  # Cellule vide sans bordure pour l'alignement
        pdf.cell(column_widths[-2], 10, "Total TTC (€)", border=1, align="R")
        pdf.cell(column_widths[-1], 10, f"{total_ttc:.2f}", border=1, align="C")



        # Enregistrer le fichier PDF
        pdf.output(fichier_pdf)
        messagebox.showinfo("Succès", f"Facture générée avec succès : {fichier_pdf}")
    
         # Ouvrir le dossier des factures dans l'explorateur
        if sys.platform == "win32":  # Pour Windows
            os.startfile(dossier_factures)
        elif sys.platform == "darwin":  # Pour macOS
            subprocess.Popen(["open", dossier_factures])
        elif sys.platform == "linux":  # Pour Linux
            subprocess.Popen(["xdg-open", dossier_factures])

    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur s'est produite : {e}")


def apercu_facture(annee, nom_proprietaire, nom_fermier, parcelles_selectionnees):
    # Création d'une fenêtre Toplevel pour l'aperçu
    apercu_window = tk.Toplevel(root)
    apercu_window.title("Aperçu de la Facture")
    apercu_window.geometry("650x700")
    apercu_window.config(bg=BACKGROUND_COLOR)

    # Charger les données du fichier Excel comme dans `generer_facture_pdf`
    try:
        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
        ws = wb[annee]
        
        adresse_proprietaire, cp_ville_proprietaire = "", ""
        adresse_fermier, cp_ville_fermier = "", ""
        prix_quintal, indice_actuel = 0, 1
        parcelle_details = []
        total_ht = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == nom_proprietaire and row[5] == nom_fermier:
                adresse_proprietaire = row[2] or "Adresse non spécifiée"
                cp_ville_proprietaire = row[3] or "CP/Ville non spécifiés"
                adresse_fermier = row[6] or "Adresse non spécifiée"
                cp_ville_fermier = row[7] or "CP/Ville non spécifiés"
                prix_quintal = float(row[10]) if row[10] else 0
                indice_actuel = float(row[11]) if row[11] else 1
                break
        
        # Calcul des détails des parcelles et totaux
        for parcelle, surface in parcelles_selectionnees:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == nom_proprietaire and row[5] == nom_fermier and row[4] == parcelle:
                    quantite = row[9] or 0
                    prix_quintal_adjusted = float(row[12]) if row[12] else 0
                    taxe = row[13] or 0
                    montant_ht = surface * quantite * prix_quintal_adjusted
                    total_ht += montant_ht
                    parcelle_details.append((parcelle, surface, quantite, prix_quintal_adjusted, taxe, montant_ht))
                    break

    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible de charger les données : {e}")
        return

    # Affichage des informations principales dans l'aperçu
    title_label = tk.Label(apercu_window, text="Aperçu de la Facture de Fermage", font=("Arial", 16, "bold"), bg=BACKGROUND_COLOR)
    title_label.pack(pady=15)
    
    # Détails propriétaire et fermier
    details_frame = tk.Frame(apercu_window, bg=BACKGROUND_COLOR, padx=20)
    details_frame.pack(pady=10, fill="x")
    details_text = f"""
    Année : {annee}
    Propriétaire : {nom_proprietaire}
    Adresse Propriétaire : {adresse_proprietaire}, {cp_ville_proprietaire}
    Fermier : {nom_fermier}
    Adresse Fermier : {adresse_fermier}, {cp_ville_fermier}
    """
    details_label = tk.Label(details_frame, text=details_text, font=("Arial", 11), justify="left", bg=BACKGROUND_COLOR, anchor="w")
    details_label.pack(anchor="w")

    # Affichage des détails des parcelles
    parcelles_frame = tk.Frame(apercu_window, bg=BACKGROUND_COLOR, relief=tk.RAISED, borderwidth=1, padx=20, pady=10)
    parcelles_frame.pack(pady=15, padx=10, fill="both", expand=True)
    tk.Label(parcelles_frame, text="Détails des parcelles :", font=("Arial", 13, "bold"), bg=BACKGROUND_COLOR).grid(row=0, column=0, columnspan=6, sticky="w", pady=(5, 5))

    # En-têtes pour les parcelles
    headers = ["Parcelle", "Surface (ha)", "Quantité (qx)", "Prix Ajusté (€)", "Impôts (%)", "Total HT (€)"]
    for col, header in enumerate(headers):
        tk.Label(parcelles_frame, text=header, font=("Arial", 10, "bold"), bg=BACKGROUND_COLOR).grid(row=1, column=col, padx=5, pady=5)

    for i, (parcelle, surface, quantite, prix_quintal_adjusted, taxe, montant_ht) in enumerate(parcelle_details):
        values = [parcelle, f"{surface:.2f}", f"{quantite:.2f}", f"{prix_quintal_adjusted:.2f}", f"{taxe:.2f}", f"{montant_ht:.2f}"]
        for col, value in enumerate(values):
            tk.Label(parcelles_frame, text=value, font=("Arial", 10), bg=BACKGROUND_COLOR).grid(row=i + 2, column=col, padx=5, pady=5)

    # Totaux HT et TTC avec style
    total_frame = tk.Frame(apercu_window, bg=BACKGROUND_COLOR, padx=20)
    total_frame.pack(pady=20, fill="x")
    total_ttc = total_ht * (1 + taxe / 100)  # Calcul du TTC avec la taxe

    tk.Label(total_frame, text=f"Total HT : {total_ht:.2f} €", font=("Arial", 12, "bold"), bg=BACKGROUND_COLOR).pack(pady=5, anchor="e")
    tk.Label(total_frame, text=f"Total TTC : {total_ttc:.2f} €", font=("Arial", 12, "bold"), bg=BACKGROUND_COLOR).pack(pady=5, anchor="e")
    
    # Boutons de confirmation, annulation et retour
    button_frame = tk.Frame(apercu_window, bg=BACKGROUND_COLOR)
    button_frame.pack(pady=15)

    tk.Button(button_frame, text="Confirmer", command=lambda: [generer_facture_pdf(annee, nom_proprietaire, nom_fermier, parcelles_selectionnees), apercu_window.destroy()], bg=BUTTON_COLOR, fg=BUTTON_TEXT_COLOR, font=("Arial", 10, "bold"), width=20).pack(side="left", padx=10)
    tk.Button(button_frame, text="Annuler", command=apercu_window.destroy, font=("Arial", 10), width=20).pack(side="right", padx=10)

    # Configuration modale pour bloquer l'accès à la fenêtre principale
    apercu_window.transient(root)
    apercu_window.grab_set()
    root.wait_window(apercu_window)


def on_generate():
    annee = combo_annee.get()
    nom_proprietaire = combo_proprietaire.get()
    nom_fermier = combo_fermier.get()
    
    parcelles_selectionnees = [
        (parcelle, float(surface))
        for parcelle, surface, var in parcelle_vars
        if var.get()
    ]

    if annee and nom_proprietaire and nom_fermier and parcelles_selectionnees:
        # Appel à la fonction d'aperçu avant génération
        apercu_facture(annee, nom_proprietaire, nom_fermier, parcelles_selectionnees)
    else:
        messagebox.showwarning("Attention", "Veuillez remplir tous les champs et sélectionner au moins une parcelle")

# Fonction pour ouvrir le lien dans le navigateur

def open_link(event):
    webbrowser.open(AUTHOR_WEBSITE)

# Interface graphique principale
root = tk.Tk()
root.title(APP_TITLE)
root.geometry(APP_GEOMETRY)
root.configure(bg=BACKGROUND_COLOR)
root.iconbitmap(ICON_PATH)

# Interface de l'application avec constantes appliquées
main_frame = tk.Frame(root, bg=BACKGROUND_COLOR)
main_frame.pack(expand=True)

# Logo
logo_image = Image.open(LOGO_PATH).resize((150, 150), Image.LANCZOS)
logo_photo = ImageTk.PhotoImage(logo_image)
logo_label = tk.Label(main_frame, image=logo_photo, bg=BACKGROUND_COLOR)
logo_label.grid(row=0, column=0, columnspan=2, pady=(20, 10))

# Boutons pour ouvrir les dossiers "factures" et "tableau" sous le logo
open_facture_button = tk.Button(main_frame, text="Ouvrir Dossier Factures", command=lambda: open_folder(PDF_OUTPUT_DIR), font=BUTTON_FONT, bg=BUTTON_COLOR, fg=BUTTON_TEXT_COLOR)
open_facture_button.grid(row=1, column=0, padx=(30, 10), pady=(5, 15), sticky="e")

open_tableau_button = tk.Button(main_frame, text="Ouvrir Dossier Tableau", command=lambda: open_folder(os.path.dirname(EXCEL_FILE_PATH)), font=BUTTON_FONT, bg=BUTTON_COLOR, fg=BUTTON_TEXT_COLOR)
open_tableau_button.grid(row=1, column=1, padx=(10, 30), pady=(5, 15), sticky="w")

# Titre
title_label = tk.Label(main_frame, text=APP_TITLE, font=TITLE_FONT, bg=BACKGROUND_COLOR)
title_label.grid(row=2, column=0, columnspan=2, pady=(0, 20))

# Champs de sélection
tk.Label(main_frame, text="Année :", bg=BACKGROUND_COLOR).grid(row=3, column=0, padx=(30, 10), pady=10, sticky="e")
combo_annee = ttk.Combobox(main_frame, values=DEFAULT_YEAR_OPTIONS, state="readonly")
combo_annee.set(DEFAULT_OPTION)  # Sélectionne l'option par défaut
combo_annee.grid(row=3, column=1, padx=(10, 30), pady=10, sticky="w")

tk.Label(main_frame, text="Nom du propriétaire :", bg=BACKGROUND_COLOR).grid(row=4, column=0, padx=(30, 10), pady=10, sticky="e")
combo_proprietaire = ttk.Combobox(main_frame, values=[], state="readonly")
combo_proprietaire.set(DEFAULT_OPTION)
combo_proprietaire.grid(row=4, column=1, padx=(10, 30), pady=10, sticky="w")

tk.Label(main_frame, text="Nom du fermier :", bg=BACKGROUND_COLOR).grid(row=5, column=0, padx=(30, 10), pady=10, sticky="e")
combo_fermier = ttk.Combobox(main_frame, values=[], state="readonly")
combo_fermier.set(DEFAULT_OPTION)
combo_fermier.grid(row=5, column=1, padx=(10, 30), pady=10, sticky="w")

# Cadre pour les parcelles
parcelles_frame = tk.Frame(main_frame, bg=BACKGROUND_COLOR)
parcelles_frame.grid(row=6, column=0, columnspan=2, padx=30, pady=10, sticky="w")
tk.Label(parcelles_frame, text="Parcelles :", bg=BACKGROUND_COLOR).grid(row=0, column=0, sticky="w")
parcelle_vars = []

# Bouton pour générer la facture
generate_button = tk.Button(main_frame, text="Générer la Facture", command=on_generate, font=BUTTON_FONT, bg=BUTTON_COLOR, fg=BUTTON_TEXT_COLOR)
generate_button.grid(row=7, column=0, columnspan=2, pady=(20, 20), padx=30, sticky="ew")

# Pied de page avec informations sur la version
footer_frame = tk.Frame(main_frame, bg=BACKGROUND_COLOR)
footer_frame.grid(row=8, column=0, columnspan=2, pady=(20, 10))

footer_text = tk.Label(footer_frame, text=f"Fermage Facturation Version {APP_VERSION}", font=LABEL_FONT, bg=BACKGROUND_COLOR)
footer_text.pack()
copyright_text = tk.Label(footer_frame, text=f"© {COPYRIGHT_YEAR} {AUTHOR_NAME}", font=LABEL_FONT, bg=BACKGROUND_COLOR)
copyright_text.pack()

link = tk.Label(footer_frame, text=AUTHOR_WEBSITE, font=LABEL_FONT, fg="blue", cursor="hand2", bg=BACKGROUND_COLOR)
link.pack()
link.bind("<Button-1>", open_link)

# Associer les événements
combo_annee.bind("<<ComboboxSelected>>", on_annee_change)
combo_proprietaire.bind("<<ComboboxSelected>>", on_proprietaire_change)

# Lancer la boucle principale
root.mainloop()
